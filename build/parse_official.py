"""[로컬 전용] 무주 재정집행 공식 종합엑셀 → data/muju_exec_official.json.
소비투자(1분기, 42통계목)·신속집행(상반기, 기초36통계목)은 다른 관리축 → 부서별 공식수치 그대로.
★ 원장으론 신속집행(큐레이션 부분집합)을 정확히 못 뽑아서 공식 스냅샷을 사용. 새 기준일 엑셀 오면 갱신.
사용: python build/parse_official.py ["종합엑셀경로"]
단위: 종합엑셀=백만원 → 원으로 변환.
"""
import os
import sys
import glob
import json

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DL = os.path.join(os.path.expanduser('~'), 'Downloads')
M = 1_000_000       # 백만원 → 원


def _n(x):
    try:
        return float(x)
    except Exception:
        return 0.0


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        cands = sorted(glob.glob(os.path.join(DL, '*재정집행_부서별_종합정리*.xlsx')), key=os.path.getmtime)
        if not cands:
            print('! 종합엑셀 없음(Downloads)'); sys.exit(1)
        path = cands[-1]
    print(f'  읽는 중: {os.path.basename(path)}')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['부서별_종합']
    # asof: 요약 시트에서
    asof = '2026-03-17'
    if '요약' in wb.sheetnames:
        for r in wb['요약'].iter_rows(values_only=True):
            for c in r:
                if c and '2026' in str(c) and '.' in str(c):
                    import re
                    m = re.search(r'2026[.\-]\s*(\d+)[.\-]\s*(\d+)', str(c))
                    if m:
                        asof = f'2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}'; break

    # 소비투자 예산현액(대상) — 소비투자_순위 시트에서 부서별 조인
    sobi_daesang = {}
    if '소비투자_순위' in wb.sheetnames:
        sw = wb['소비투자_순위']
        for r in sw.iter_rows(min_row=2, values_only=True):
            if r[1]:            # [순위, 부서, 예산현액_계, ...]
                sobi_daesang[str(r[1]).strip()] = round(_n(r[2]) * M)

    sobi, sok = [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0] or str(r[0]).strip() in ('계', '합계', '총계'):
            continue
        name = str(r[0]).strip()
        sobi.append({'name': name, 'target': sobi_daesang.get(name, 0), 'goal': round(_n(r[2]) * M),
                     'exec': round(_n(r[3]) * M), 'rate': round(_n(r[4]) * 100, 1)})     # 소비투자: 대상(현액)/목표/집행/집행률
        sok.append({'name': name, 'target': round(_n(r[9]) * M), 'goal': round(_n(r[10]) * M),
                    'exec': round(_n(r[11]) * M), 'rate': round(_n(r[12]) * 100, 1)})    # 신속집행: 대상/목표(60%)/집행/목표대비율

    def tot(rows, keys):
        return {k: sum(x[k] for x in rows) for k in keys}
    st = tot(sobi, ('target', 'goal', 'exec')); st['rate'] = round(st['exec'] / st['goal'] * 100, 1) if st['goal'] else 0
    st['trate'] = round(st['goal'] / st['target'] * 100, 1) if st['target'] else 0     # 대상 대비 목표율(참조선)
    kt = tot(sok, ('target', 'goal', 'exec')); kt['rate'] = round(kt['exec'] / kt['goal'] * 100, 1) if kt['goal'] else 0
    kt['trate'] = round(kt['goal'] / kt['target'] * 100, 1) if kt['target'] else 0     # 60%

    out = {'asof': asof,
           'sobi': {'label': '소비투자', 'period': '1분기', 'depts': sobi, 'total': st,
                    'note': '1·2분기 경기보강 · 42통계목'},
           'sok': {'label': '신속집행', 'period': '상반기', 'depts': sok, 'total': kt,
                   'note': '상반기 조기집행 · 기초36통계목 · 목표=대상의 60%'},
           'source': f'무주군 기획조정실 재정집행 종합({asof} 기준)'}
    dst = os.path.join(ROOT, 'data', 'muju_exec_official.json')
    json.dump(out, open(dst, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
    print(f"✓ data/muju_exec_official.json — {asof} · 소비투자 {len(sobi)}부서(집행률 {st['rate']}%) · 신속집행 {len(sok)}부서(목표대비 {kt['rate']}%)")


if __name__ == '__main__':
    main()
